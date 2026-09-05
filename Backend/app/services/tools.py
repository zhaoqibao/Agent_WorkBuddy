"""工具体系：langchain @tool 写法，供 create_deep_agent 使用。

每个工具都用 @tool 装饰器 + 详细 docstring（描述用途与参数含义），
确保智能体（agent）能根据描述正确识别并调用。

数据隔离：需要访问数据库的工具通过 `config` 注入 user_id，
工具内部用独立的数据库会话查询，只返回当前用户的数据。
"""
from __future__ import annotations

import base64
import csv
import io
import json
import datetime

import httpx
from langchain_core.messages import HumanMessage
from langchain_core.runnables import RunnableConfig
from langchain_core.tools import tool

from app.core.config import settings
from app.core.database import SessionLocal
from app.models import Document
from app.services.llm import get_vlm

# 图片类型：read_document 遇到这些类型时走视觉模型识别
_IMAGE_TYPES = {"png", "jpg", "jpeg", "gif", "webp", "bmp", "svg"}


def _uid(config: RunnableConfig) -> int | None:
    return (config or {}).get("configurable", {}).get("user_id")


async def _vlm_describe(data_url: str, prompt: str) -> str:
    """用视觉模型识别图片内容。"""
    vlm = get_vlm()
    msg = HumanMessage(
        content=[
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": data_url}},
        ]
    )
    try:
        resp = await vlm.ainvoke([msg])
        return resp.content or ""
    except Exception as e:
        return f"图片识别失败: {e}"


# ---------- 工具 0：获取时间 ----------
@tool
def get_current_time():
  """
  获取当前时间
  当用户询问内容涉及时间时调用此工具。
  """
  current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  return current_time


# ---------- 工具 1：天气查询 ----------
@tool
async def get_weather(city: str) -> str:
    """查询指定城市的实时天气，返回温度、风速和天气状况。

    当用户询问某个城市的天气时调用此工具。

    Args:
        city: 城市名称，例如「北京」「上海」「深圳」
    """
    if not city:
        return "未提供城市名"
    async with httpx.AsyncClient(timeout=20.0) as client:
        geo = await client.get(
            "https://geocoding-api.open-meteo.com/v1/search",
            params={"name": city, "language": "zh", "count": 1},
        )
        geo.raise_for_status()
        results = geo.json().get("results") or []
        if not results:
            return f"未找到城市「{city}」"
        loc = results[0]
        lat, lon = loc["latitude"], loc["longitude"]
        name = loc.get("name", city)
        w = await client.get(
            "https://api.open-meteo.com/v1/forecast",
            params={"latitude": lat, "longitude": lon, "current_weather": True},
        )
        w.raise_for_status()
        cw = w.json().get("current_weather") or {}
        return (
            f"{name}当前天气：温度 {cw.get('temperature')}°C，"
            f"风速 {cw.get('windspeed')} km/h，"
            f"天气代码 {cw.get('weathercode')}（0 晴/1-3 多云/45-48 雾/51-67 雨/71-77 雪/95-99 雷暴）"
        )


# ---------- 工具 2：新闻搜索 ----------
@tool
async def get_news(query: str) -> str:
    """搜索实时新闻或资讯（基于 Tavily 搜索），返回最新结果的标题和链接。

    当用户需要了解最新消息、新闻、时事或某个话题的实时信息时调用。

    Args:
        query: 搜索关键词或问题，例如「人工智能最新进展」
    """
    if not query:
        return "未提供搜索关键词"
    if not settings.TAVILY_SEARCH_KEY:
        return "未配置 TAVILY_SEARCH_KEY，新闻搜索不可用"
    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.post(
            "https://api.tavily.com/search",
            json={
                "api_key": settings.TAVILY_SEARCH_KEY,
                "query": query,
                "max_results": 5,
                "search_depth": "basic",
            },
        )
        resp.raise_for_status()
        data = resp.json()
    items = data.get("results") or []
    if not items:
        return f"未找到与「{query}」相关的新闻"
    lines = [f"关于「{query}」的最新结果："]
    for it in items:
        lines.append(f"- {it.get('title', '')}（{it.get('url', '')}）")
    return "\n".join(lines)


# ---------- 工具 3：读取文档（含图片识别） ----------
async def _recognize_document_image(doc: Document) -> str:
    """图片文档：从 MinIO 下载后用视觉模型识别内容。"""
    from app.services import storage

    try:
        raw = storage.get_object(doc.stored_path)
    except Exception as e:
        return f"读取图片失败: {e}"

    ext = (doc.file_type or "").lower()
    mime = "jpeg" if ext in ("jpg", "jpeg") else (ext or "png")
    data_url = f"data:image/{mime};base64,{base64.b64encode(raw).decode()}"
    if not settings.BASE_VLM:
        return "未配置视觉模型（BASE_VLM），无法识别图片"
    return await _vlm_describe(data_url, "请详细描述这张图片的内容")


@tool
async def read_document(document_id: int, config: RunnableConfig) -> str:
    """读取资料库中某个文档的内容。

    当用户提到某个已上传的文档/文件（如「总结一下这个文档」「读一下这份文件」），
    且提供了 document_id 时调用。若文档是图片，会自动用视觉模型识别并描述图片内容。

    Args:
        document_id: 文档的 ID（整数，来自用户消息中提到的 document_id）
    """
    uid = _uid(config)
    async with SessionLocal() as db:
        doc = await db.get(Document, document_id)
        if not doc or doc.user_id != uid:
            return "文档不存在或无权访问"

        if (doc.file_type or "").lower() in _IMAGE_TYPES:
            return await _recognize_document_image(doc)

        text = doc.text_content or ""
        if not text:
            return "该文档暂无解析文本"
        return text[:6000] if len(text) > 6000 else text


# ---------- 工具 4：文档格式转换 ----------
def _extract_text(filename: str, data: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext in ("txt", "md", "csv"):
            return data.decode("utf-8", errors="ignore")
        if ext == "docx":
            from docx import Document as Docx
            return "\n".join(p.text for p in Docx(io.BytesIO(data)).paragraphs)
        if ext == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    out.append("\t".join("" if c is None else str(c) for c in row))
            return "\n".join(out)
        if ext == "pdf":
            import pdfplumber
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                return "\n".join((p.extract_text() or "") for p in pdf.pages)
    except Exception as e:
        return f"[解析失败: {e}]"
    return "[不支持的格式]"


def _xlsx_to_csv(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    buf = io.StringIO()
    writer = csv.writer(buf)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in row])
    return buf.getvalue()


def _csv_to_json(text: str) -> str:
    reader = csv.DictReader(io.StringIO(text))
    rows = [row for row in reader]
    return json.dumps(rows, ensure_ascii=False, indent=2)


@tool
async def convert_document(document_id: int, target_format: str, config: RunnableConfig):
    """把资料库中的文档转换为指定格式，结果保存为文件供用户下载（不在对话里展示全部内容）。

    当用户要求把文档转成 txt/md/csv/json 等格式时调用。

    Args:
        document_id: 文档的 ID（整数）
        target_format: 目标格式，支持 txt、md、csv、json
    """
    import uuid

    uid = _uid(config)
    from app.services import storage

    async with SessionLocal() as db:
        doc = await db.get(Document, document_id)
        if not doc or doc.user_id != uid:
            return "文档不存在或无权访问"

        try:
            raw = storage.get_object(doc.stored_path)
        except Exception as e:
            return f"读取文档失败: {e}"

        target = (target_format or "txt").lower().strip(".")
        text = _extract_text(doc.original_name, raw)

        if target in ("txt", "md"):
            result = text
        elif target == "csv":
            result = _xlsx_to_csv(raw) if (doc.file_type or "") == "xlsx" else text
        elif target == "json":
            result = _csv_to_json(text) if (doc.file_type or "") == "csv" else json.dumps({"content": text}, ensure_ascii=False, indent=2)
        else:
            return f"暂不支持转换到 {target}（支持 txt/md/csv/json）"

        # 转换结果保存到 MinIO，返回对象 key 供后端下载接口使用（避免把大段内容直接塞进对话）
        base = (doc.original_name or "document").rsplit(".", 1)[0]
        filename = f"{base}.{target}"
        key = f"converted/u-{uid}/{uuid.uuid4().hex}.{target}"
        data = result.encode("utf-8")
        storage.put_object(key, data, len(data), "text/plain; charset=utf-8")

        # 预览：仅取前两行非空内容
        lines = [ln for ln in result.splitlines() if ln.strip()][:2]
        preview = "\n".join(lines)[:200]

        # 返回 JSON（含对象 key + 文件名 + 预览），由后端解析后传给前端展示下载按钮
        info = {
            "message": f"已将「{doc.original_name}」转换为 {target} 格式（共 {len(result)} 字符），可点击下载。",
            "key": key,
            "filename": filename,
            "preview": preview,
        }
        return json.dumps(info, ensure_ascii=False)


# ---------- 工具 5：图片识别 ----------
def _minio_endpoint_host() -> str:
    """MinIO endpoint 的 host 部分（去掉 scheme、端口、路径）。"""
    ep = (settings.MINIO_ENDPOINT or "").strip()
    for scheme in ("http://", "https://"):
        if ep.startswith(scheme):
            ep = ep[len(scheme):]
    return ep.split("/")[0].split(":")[0]


def _image_url_to_data_url(image_url: str) -> str:
    """把图片地址统一解析为可被视觉模型访问的地址。

    - data URL（data:image/...）原样返回；
    - MinIO 对象 key（无 http 前缀，如 ws-global/u-1/xxx.png）→ 从 MinIO 下载转 data URL；
    - MinIO 预签名 URL → 从 URL 提取对象 key 后下载转 data URL；
    - 其他外部 http(s) URL → 原样返回（交给视觉模型直接访问）。
    """
    import base64
    from urllib.parse import urlparse

    if image_url.startswith("data:"):
        return image_url

    from app.services import storage

    key = None
    if image_url.startswith(("http://", "https://")):
        parsed = urlparse(image_url)
        if parsed.netloc.split(":")[0] == _minio_endpoint_host():
            path = parsed.path.lstrip("/")
            if path.startswith(settings.MINIO_BUCKET + "/"):
                path = path[len(settings.MINIO_BUCKET) + 1:]
            key = path
        else:
            return image_url  # 外部 URL，直接交给视觉模型
    else:
        key = image_url  # 视为 MinIO 对象 key

    if not key:
        return image_url

    try:
        raw = storage.get_object(key)
    except Exception as e:
        return f"读取图片失败: {e}"

    ext = key.rsplit(".", 1)[-1].lower() if "." in key else "png"
    mime = "jpeg" if ext in ("jpg", "jpeg") else (ext or "png")
    return f"data:image/{mime};base64,{base64.b64encode(raw).decode()}"


@tool
async def recognize_image(image_url: str, prompt: str = "请描述这张图片的内容") -> str:
    """识别或理解一张图片的内容（基于视觉大模型）。

    当用户提供图片地址并询问图片内容时调用。图片地址可以是外部 URL，
    也可以是资料库/MinIO 中的图片（此时会先下载再用视觉模型识别）。

    Args:
        image_url: 图片的 URL 地址（http/https 或 data URL，或 MinIO 对象路径）
        prompt: 关于图片的问题，例如「描述这张图」「图里有什么」
    """
    if not image_url:
        return "未提供图片地址（image_url）"
    if not settings.BASE_VLM:
        return "未配置视觉模型（BASE_VLM）"
    data_url = _image_url_to_data_url(image_url)
    return await _vlm_describe(data_url, prompt)


# ---------- 工具 6：图片生成 ----------
@tool
async def generate_image(prompt: str, size: str = "1024x1024", config: RunnableConfig = None):
    """根据文字描述生成一张图片，生成结果会直接展示在对话中（并保存到对象存储，刷新后仍可查看/下载）。

    当用户要求画图、生成图片、创作图像时调用。

    Args:
        prompt: 图片的文字描述，尽量详细
        size: 图片尺寸，如 1024x1024、1328x1328 等
    """
    import base64
    import uuid

    if not prompt:
        return "未提供图片描述（prompt）"
    if not settings.IMAGE_MODEL:
        return "未配置图片生成模型（IMAGE_MODEL）"

    payload = {"model": settings.IMAGE_MODEL, "prompt": prompt, "size": size, "n": 1}
    headers = {"Authorization": f"Bearer {settings.OPENAI_API_KEY}"}
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(
            f"{settings.MODEL_API_BASE_URL.rstrip('/')}/images/generations",
            json=payload,
            headers=headers,
        )
        resp.raise_for_status()
        data = resp.json()
    try:
        item = data["data"][0]
        b64 = item.get("b64_json")
        url = item.get("url")
    except (KeyError, IndexError):
        return "图片生成失败"

    # 下载图片字节（base64 或 url）
    img_bytes = None
    if b64:
        img_bytes = base64.b64decode(b64)
    elif url:
        async with httpx.AsyncClient(timeout=60.0) as client:
            r = await client.get(url)
            r.raise_for_status()
            img_bytes = r.content
    if not img_bytes:
        return "图片生成失败"

    # 保存到 MinIO（持久，刷新后可恢复）
    from app.services import storage

    uid = _uid(config)
    key = f"generated/u-{uid}/{uuid.uuid4().hex}.png"
    storage.put_object(key, img_bytes, len(img_bytes), "image/png")

    info = {
        "message": "已生成图片，展示在对话中。",
        "image_key": key,
    }
    return json.dumps(info, ensure_ascii=False)


# ---------- 工具集合 ----------
def get_agent_tools() -> list:
    """返回智能体可用的全部工具列表（供 create_deep_agent 使用）。"""
    return [
        get_weather,
        get_news,
        read_document,
        convert_document,
        recognize_image,
        generate_image,
        get_current_time
    ]
